"""
    An AI Tool for Student-Supervisor Allocation.
    
    Package: pystsup
    Module: utilities
    File: generateData.py
    
    Purpose:  Contains Required functions to read input data from excel file, save results into an excel file
              and create input data excel file.
             
    Author : Rithin Chalumuri
    Version: 1.0 
    Date   : 21/7/17
    
"""

import random
import openpyxl as xl
from openpyxl import *
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.chart import *

from .acmParser import getPath, parseFile
from ga.pystsup.data import Student, Supervisor

def getRandomKeyword(keyword, supNames, keywordsFile):
    topicNames,topicPaths,topicIDs,levels = parseFile(keywordsFile)
    keylist = []
    keyword = keyword.lower()
    topicName = topicNames[keyword]
    for i in range(1, len(levels)+1):
        lent = getPath(topicIDs[i], topicNames, topicPaths, topicIDs, supNames)
        if levels[i] > levels[topicName] and lent[len(lent)-2] == keyword:
            keylist.append(topicIDs[i])
    randomKey = random.choice(keylist)
    randomKey2 = random.choice(keylist)
    return randomKey, randomKey2

def getData(stuFile,supFile, keywordsFile):

    topicNames,topicPaths,topicIDs,levels = parseFile(keywordsFile)

    m = len(stuFile)
    n = len(supFile)

    numberOfStudents = n
    numberOfSupervisors = m

    stuID = "student"
    supID = "supervisor"

    students = {}
    supervisors = {}

    #Getting Students
    
    for row in stuFile:
        
        n-=1
        rank = 0
        stu_List = {}
        
        stu = stuID + str(numberOfStudents - n)
        
        realID = row[0]
        name = row[1] + " " + row[2] + ". " + row[3] + "."
        supName = [row[6], row[7]]
        keywords = []
        keywords.extend(getRandomKeyword(row[4], supName, keywordsFile))
        keywords.extend(getRandomKeyword(row[5], supName, keywordsFile))
        
                
        for kw in keywords:
            rank+=1
            stu_List[rank] = [kw,getPath(kw,topicNames, topicPaths, topicIDs, supName)]
        
        studentObject = Student(stu,stu_List,realID,name)
        students[stu] = studentObject


    #Getting Supervisors
    
    for row in supFile:

        m-=1
        rank = 0
        sup_List = {}

        sup = supID + str(numberOfSupervisors - m)

        realID = row[0]
        name = row[1] + " " + row[2] + ". " + row[3] + "."
        quota = row[4]
        keywords = []
        supName = [row[1], row[1]]
        keywords.extend(getRandomKeyword(row[5], supName, keywordsFile))
        keywords.extend(getRandomKeyword(row[6], supName, keywordsFile))

        for kw in keywords:
            rank+=1
            sup_List[rank] = [kw,getPath(kw,topicNames, topicPaths, topicIDs, row[1])]

        supervisorObject = Supervisor(sup,sup_List,quota,realID,name)
        supervisors[sup] = supervisorObject
        

    return students,supervisors

def writeFrontier(filename, front,metricData, supervisors, students):
    """
    Function to save the final results in an excel file after the completion of the GA.

    Parameters:
    
        filename (string) - the name of the file we want save results in.
        front (list) - list of solutions in the first front.
        metricData (dictionary) - metric data associated with the GA (time taken, max Fitness, diversity etc.)
        supervisors (dictionary) - dictionary of all supervisors with their details (id,preferences,quota, name, realID)
        students (dictionary) - dictionary of all students with their details (id, preferences ,name, realID)
    """

    filename += '.xlsx'
    
    wb = Workbook()

    #Creating the Statistics Sheet
    
    ws = wb.active
    ws.title = "GA Statistics"
    ws.append(("Total Number of Generations",metricData['numberOfGenerations']))
    ws.append(("Total Time Taken (s)",metricData['total_time_generations']))
    ws.append(("Time Per Generation (s)",metricData['avg_time_generation']))
    ws.append((None,None))
    ws.append(("Initial Maximum Student Fitness (Fst)",metricData['initial_maxFst']))
    ws.append(("Final Maximum Student Fitness (Fst)",metricData['final_maxFst']))
    ws.append((None,None))
    ws.append(("Initial Minimum Student Fitness (Fst)",metricData['initial_minFst']))
    ws.append(("Final Minimum Student Fitness (Fst)",metricData['final_minFst']))
    ws.append((None,None))
    ws.append(("Initial Maximum Supervisor Fitness (Fsup)",metricData['initial_maxFsup']))
    ws.append(("Final Maximum Supervisor Fitness (Fsup)",metricData['final_maxFsup']))
    ws.append((None,None))
    ws.append(("Initial Minimum Supervisor Fitness (Fsup)",metricData['initial_minFsup']))
    ws.append(("Final Minimum Supervisor Fitness (Fsup)",metricData['final_minFsup']))
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 17

    
    #Creating the Overview Sheet
    wb.create_sheet("Overview")
    ws = wb["Overview"]
     
    heading = ("Sheet Number","Student Fitness","Supervisor Fitness")
    ws.append(heading)
    
    
    maxFst = max(front, key=lambda x: x.getFst()).getFst()
    maxFsup = max(front, key=lambda x: x.getFsup()).getFsup()
    minFst = min(front,key=lambda x: x.getFst()).getFst()
    minFsup = min(front,key=lambda x: x.getFsup()).getFsup()


    #Filter the frontier with only unique values

    front = set(front)
    n = len(front)
    
    if len(front) > 1:
        front = sorted(front,key=lambda x: (((x.getFsup()-minFsup)/(maxFsup-minFsup))*((x.getFst()-minFst)/(maxFst-minFst))),reverse=True)
        
        
    
    count = 1

    sheetnames = []
    widths = [25,25,15,15]
    letters = ['A','B','C','D']
    
    for sol in front:
        
        sheetName = "Solution " + str(count)
        wb.create_sheet(sheetName)
        sheetnames.append((sheetName,sol.getFst(),sol.getFsup()))
        ws = wb[sheetName]
        

        supEdges = sol.getGraph().getEdges()

        heading = ("Supervisor Name","Student Name","Supervisor ID","Student ID")
        
        ws.append(heading)
        
        for sup in supEdges:
            
            supName = supervisors[sup].getSupervisorName()
            supID = supervisors[sup].getRealID()
            
            for stu in supEdges[sup]:

                stuName = students[stu].getStudentName()
                stuID = students[stu].getRealID()

                tempRow = (supName,stuName,supID,stuID)
                
                ws.append(tempRow)


        for i,width in enumerate(widths):
            ws.column_dimensions[letters[i]].width = width
                

        count+=1


    #Change current worksheet to 'Overview'
        
    ws = wb['Overview']

    #Add the Data into cells
    
    for i in sheetnames:
        ws.append(i)

    ws["A"+str(n+4)]="Diversity Metric"
    ws["B"+str(n+4)] = metricData['diversity']

    #Create the chart and add it to the worksheet
    
    chart = ScatterChart(scatterStyle='marker')
    chart.title = "Pareto Optimal Frontier"
    chart.x_axis.title = 'Student Fitness (Fst)'
    chart.y_axis.title = 'Supervisor Fitness (Fsup)'
    chart.style = 3
    
    xvalues = Reference(ws,min_col=2,min_row=2,max_row=n+1)

    yvalues = Reference(ws,min_col=3,min_row=2,max_row=n+1)

    series = Series(yvalues,xvalues)

    series.marker = xl.chart.marker.Marker('circle')
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    chart.legend = None
    
    ws.add_chart(chart,"A"+str(n+7))

    #Set the column Widths
    
    ws.column_dimensions['A'].width = 17
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 17

    #Save the File
    
    wb.save(filename)

    
                   
        
